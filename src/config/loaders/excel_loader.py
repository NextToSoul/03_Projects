"""
Excel 协议表加载器
从 Excel 遥测表、指令表、注入表中读取数据，构造运行时 data model
"""

from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ...core.models import (
    TelemetryParam, TelemetryPackageDef,
    CommandDef, InjectionParamDef, InjectionDef,
)


class ExcelLoader:
    """Excel 协议表加载器"""

    @staticmethod
    def load_telemetry_packages(excel_path: str | Path) -> list[TelemetryPackageDef]:
        """
        加载遥测配置表.xlsx
        格式：每个 Sheet = 一个遥测包；每行 = 一个参数
        标题行结构：序号 | 参数ID | 参数名称 | data_offset | bit_offset
                   | bit_length | type | endian | scale | decimal_places
                   | unit | range_min | range_max | 枚举取值说明
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required to load Excel files")
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        packages = []

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("~$"):
                continue  # 跳过临时文件
            
            ws = wb[sheet_name]
            params = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                param_id = row[1]  # B列: 参数ID
                if param_id is None:
                    continue
                
                param = TelemetryParam(
                    id=str(param_id).strip(),
                    name=str(row[2]).strip() if row[2] else "",
                    data_offset=int(row[3]) if row[3] is not None else 0,
                    bit_offset=int(row[4]) if row[4] is not None else 0,
                    bit_length=int(row[5]) if row[5] is not None else 8,
                    data_type=str(row[6]).strip().lower() if row[6] else "uint8",
                    endian=str(row[7]).strip().lower() if row[7] else "big",
                    scale=float(row[8]) if row[8] is not None else 1.0,
                    decimal_places=int(row[9]) if row[9] is not None else None,
                    unit=str(row[10]).strip() if row[10] else "",
                    range_min=ExcelLoader._try_float(row[11]),
                    range_max=ExcelLoader._try_float(row[12]),
                    enum_values=ExcelLoader._parse_enum(
                        str(row[13]) if row[13] else ""
                    ),
                )
                params.append(param)

            packages.append(TelemetryPackageDef(
                name=sheet_name,
                parameters=params,
            ))

        return packages

    @staticmethod
    def load_commands(excel_path: str | Path) -> list[CommandDef]:
        """
        加载立即遥控指令配置表.xlsx
        标题行结构：序号 | 指令代号 | 指令名称 | 帧头标识位
                   | 应用过程标识 | 指令长度 | 指令编码 | 参数值 | 是否轮询
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required to load Excel files")
        
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        commands = []

        consecutive_empty = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cmd_id = row[1]  # B列: 指令代号
            if cmd_id is None:
                consecutive_empty += 1
                if consecutive_empty > 1000:  # 连续1000行空数据则停止
                    break
                continue
            consecutive_empty = 0
            cmd_name = str(row[2]).strip() if row[2] else ""
            
            commands.append(CommandDef(
                id=str(cmd_id).strip(),
                name=cmd_name,
                frame_header=str(row[3]).strip() if row[3] else "EB 90",
                app_process_id=str(row[4]).strip() if row[4] else "05 20",
                data_length=str(row[5]).strip() if row[5] else "00 01",
                command_code=str(row[6]).strip() if row[6] else "",
                default_param=str(row[7]).strip() if row[7] else "",
                is_polling=str(row[8]).strip() == "是" if row[8] else False,
            ))

        return commands

    @staticmethod
    def load_injections(excel_paths: list[dict]) -> list[InjectionDef]:
        """加载固定地址参数注入表"""
        injections = []
        for inj_def in excel_paths:
            path = inj_def.get("excel_path", "")
            name = inj_def.get("name", "Unknown")
            if not path:
                continue
            
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Parameters"]
            params = []
            
            for row in ws.iter_rows(min_row=4, values_only=True):
                # Row format: No. | English Name | Chinese Name | Byte Offset
                #              | Byte Length | Unit | Data Type | Default | Remark
                chinese_name = row[2]
                if chinese_name is None:
                    continue
                
                params.append(InjectionParamDef(
                    name=str(chinese_name).strip(),
                    byte_offset=int(row[3]) if row[3] is not None else 0,
                    byte_length=int(row[4]) if row[4] is not None else 4,
                    data_type=str(row[6]).strip().lower() if row[6] else "float32",
                    default_value=str(row[7]).strip() if row[7] else "",
                    unit=str(row[5]).strip() if row[5] else "",
                ))
            
            injections.append(InjectionDef(
                name=name,
                parameters=params,
            ))
        
        return injections



    @staticmethod
    def _try_float(value) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None



    @staticmethod
    def _try_float(value) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_enum(enum_str: str) -> dict[int, str] | None:
        """解析枚举列：'0:待机模式; 1:推进模式; 2:应急模式'"""
        if not enum_str or enum_str == "None":
            return None
        result = {}
        for pair in enum_str.split(";"):
            pair = pair.strip()
            if ":" in pair:
                key, val = pair.split(":", 1)
                try:
                    result[int(key.strip())] = val.strip()
                except ValueError:
                    continue
        return result if result else None
